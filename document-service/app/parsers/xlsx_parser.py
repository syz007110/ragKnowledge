from __future__ import annotations

from io import BytesIO

from app.parsers.types import BaseParser, IngestParseResult
from app.parsers.unified_builders import (
    IdGen,
    assemble_parse_document,
    legacy_xlsx_payload,
    raw_text_from_pages,
    table_block_nested,
    title_block,
)

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


def _row_to_kv(headers: list[str], values: list[str]) -> str:
    pieces: list[str] = []
    max_len = max(len(headers), len(values))
    for idx in range(max_len):
        header = (headers[idx] if idx < len(headers) else "") or f"列{idx + 1}"
        value = values[idx] if idx < len(values) else ""
        value = str(value or "").strip()
        if not value:
            continue
        pieces.append(f"{header}: {value}")
    return "; ".join(pieces)


class XlsxParser(BaseParser):
    name = "xlsx_openpyxl"

    def parse(
        self,
        file_bytes: bytes,
        *,
        filename: str,
        file_ext: str,
        mime_type: str,
    ) -> IngestParseResult:
        if not load_workbook:
            raise RuntimeError("parser.xlsxUnavailable")
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        id_gen = IdGen("u")
        blocks: list[dict] = []
        ro = 0

        def next_ro() -> int:
            nonlocal ro
            v = ro
            ro += 1
            return v

        text_rows: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[str(cell or "").strip() for cell in row] for row in ws.iter_rows(values_only=True)]
            compact_rows = [row for row in rows if any(row)]
            if not compact_rows:
                continue
            st = title_block(id_gen, f"工作表 {sheet_name}", 1, next_ro())
            if st:
                blocks.append(st)
            tb = table_block_nested(id_gen, compact_rows, next_ro(), caption=None)
            if tb:
                blocks.append(tb)
            header = compact_rows[0]
            for idx, row in enumerate(compact_rows[1:], start=1):
                row_kv = _row_to_kv(header, row)
                if row_kv:
                    text_rows.append(f"[{sheet_name}] {row_kv}")

        pages = [{"pageIndex": 0, "blocks": blocks}]
        parser_kind = "xlsx_openpyxl_v1"
        doc = assemble_parse_document(
            file_ext="xlsx",
            parser_kind=parser_kind,
            source_file_name=filename or "",
            pages=pages,
            assets=[],
            warnings=[],
        )
        raw = raw_text_from_pages(pages)
        if not raw.strip():
            raw = "\n".join(text_rows)
        legacy = legacy_xlsx_payload(doc)
        return IngestParseResult(
            raw_text=raw,
            xlsx=legacy,
            parse_document=doc,
        )
